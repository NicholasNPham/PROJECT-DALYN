# STANDARD LIBRARY IMPORTS
from pathlib import Path

# LOCAL FILE IMPORTS
from models import ClassificationResult

# THIRD-PARTY IMPORTS
from openpyxl import load_workbook

def parse_phrases(excel_path: Path) -> dict[tuple[str, str], list[str]]:
    """
    Reads the DALYN Excel phrase sheet and builds a scoring dictionary.

    Opens the workbook at the given path, iterates over every data row,
    and groups phrases by their (type, subtype) pair. The Document Name
    column is ignored. Each phrase is stored as-is; lowercasing happens
    at match time in score_document.

    Args:
        excel_path: Path to the DALYN.xlsx file.

    Returns:
        A dictionary mapping (document_type, document_subtype) tuples
        to a list of phrases belonging to that classification pair.

    Raises:
        FileNotFoundError: If the file at excel_path does not exist.
        InvalidFileException: If the file is not a valid Excel workbook.
    """

    # LOCAL CONSTANTS
    col_phrase = 1 # in the Excel, column b is the phrases
    col_type = 2 # in the Excel, column c is the type
    col_subtype = 3 # in the Excel, column d is the subtype

    phrase_dict = {} # init an empty dictionary

    workbook = load_workbook(excel_path) # load the Excel sheet
    worksheet = workbook.active # make the worksheet in the workbook active

    for row in worksheet.iter_rows(min_row=2): # loop through every row and skip the header row
        phrase = row[col_phrase].value # init phrase_column and store the value in the cell for that row
        doc_type = row[col_type].value # init type_column and store the value in the cell for that row
        doc_subtype = row[col_subtype].value # init subtype_column and store the value in the cell for that row

        if phrase is None or doc_type is None or doc_subtype is None: # if all three are none
            continue # skip it by continuing it

        key = (doc_type, doc_subtype) # init a tuple with type and subtype column values
        if key not in phrase_dict: # if key is not in the phrase_dict dictionary as a key
            phrase_dict[key] = [] # create a new key, (empty)value pair
        phrase_dict[key].append(phrase) # if it is found then append the phrase column to the existing key

    return phrase_dict

def score_document(phrase_dict: dict[tuple[str, str], list[str]], whole_pdf_text: str) -> ClassificationResult:
    """
    Scores extracted PDF text against the phrase dictionary and returns a classification.

    Iterates over every (type, subtype) pair and its associated phrases, checking
    whether each phrase appears in the extracted text. Each match adds one point to
    that pair's cumulative score. The pair with the highest score wins and populates
    the returned ClassificationResult. Ties and zero matches result in a ClassificationResult
    with no type or subtype set, signaling that the document requires manual review.

    Both the extracted text and each phrase are lowercased before comparison to ensure
    case-insensitive matching.

    Args:
        phrase_dict: A dictionary mapping (document_type, document_subtype) tuples
                     to a list of phrases, as returned by parse_phrases.
        whole_pdf_text: The full extracted text of the PDF document as a plain string.

    Returns:
        A ClassificationResult populated with the winning type, subtype, score,
        and matched phrases. If no phrases matched or a tie occurred, document_type
        and document_subtype will be None.
    """

    score_dict = {} # init score dictionary
    matched_dict = {} # init match phrases dictionary

    for type_and_subtype, phrases in phrase_dict.items(): # separate the type and subtype tuple and the phrases and init them
        for phrase in phrases: # loop through phrases list
            if phrase.lower() in whole_pdf_text.lower(): # check if phrase is in the whole pdf txt and lowercase

                if type_and_subtype not in score_dict: # if the key is not in score dict
                    score_dict[type_and_subtype] = 0 # create a new key value pair with value set to 0
                score_dict[type_and_subtype] += 1 # add 1 to the key value pair if it exists

                if type_and_subtype not in matched_dict: # if the key not in the matched dict
                    matched_dict[type_and_subtype] = [] # create a new key value pair with a value set to an empty list
                matched_dict[type_and_subtype].append(phrase) # append the phrase string to the list

    if not score_dict: # if score dict is empty meaning no phrases were found and cant be scored
        return ClassificationResult() # return an empty dataclass

    """
    max(score_dict, ..., ...): Iterates through the score_dict dictionary and returns the keys in the dictionary.

    Note:
    - LAMBDA: lambda is a tiny, throwaway function written inline, with no def and no name.
    	
    	def get_score(key):
    		return score_dict[key]
    	
    	is the same as:
    	
    	lambda key: score_dict[key]

    This line determines what each key has the highest value then stores the key(tuple) associated with the highest value
    """ # This is for the line below to understand
    winner_type_and_subtype = max(score_dict, key=lambda key: score_dict[key]) # look for the highest point and save the winner type and subtype this finds the first winner NOTE: returns the key/tuple of the winner

    winning_score = score_dict[winner_type_and_subtype] # store the winning score example: int/2
    tie_count = 0 # init tie count variable
    for score in score_dict.values(): # loop through the score dict values
        if score == winning_score: # if we find the score that matches the winning score
            tie_count += 1 # add to the tie count list

    if tie_count > 1: # if two types and subtype key wins
        return ClassificationResult() # return an empty dataclass

    # if only one wins, return a winning classification result.
    return ClassificationResult(
        winner_type_and_subtype[0], # parameter of type
        winner_type_and_subtype[1], # parameter of subtype
        score_dict[winner_type_and_subtype], # parameter of the int score
        matched_dict[winner_type_and_subtype]) # parameter of the whole phrase list